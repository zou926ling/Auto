# 读取全局变量sheet工作表
import re

import openpyxl
from jsonpath_rw import Index, Fields
from jsonpath_rw_ext import parse

from apiddtframework1.common import util_func
from apiddtframework1.common.util_func import cur_timestamp


def get_variables(wb):
    sheet_data = wb['全局变量']
    variables = {}  # 用来存储读到的变量，名称是key，值是value
    lines_count = sheet_data.max_row  # 获取总行数
    for l in range(2, lines_count + 1):
        key = sheet_data.cell(l, 1).value
        value = sheet_data.cell(l, 2).value
        variables[key] = value
    # 增加一个内置变量，叫时间戳，注意这个时间戳是当前测试一运行就会产生，产生之后在当前测试未完成之前不管调用
    # 多少次，都是一致的
    variables['timestamp'] = cur_timestamp()
    return variables


def get_api_default_params(wb):
    sheet_data = wb['接口默认参数']
    api_default_params = {}  # 用来存储读到的变量，名称是key，值是value
    lines_count = sheet_data.max_row  # 获取总行数
    for l in range(2, lines_count + 1):
        key = sheet_data.cell(l, 1).value
        value = sheet_data.cell(l, 2).value
        api_default_params[key] = value
    return api_default_params


# 获取要执行的测试集合名称
def get_casesuitename(wb):
    sheet_data = wb['测试集合管理']
    lines_count = sheet_data.max_row  # 获取总行数
    cases_suite_name = []  # 用来存储要执行的测试集合名称
    for l in range(2, lines_count + 1):
        flag = sheet_data.cell(l, 2).value
        if flag == 'y':
            suite_name = sheet_data.cell(l, 1).value
            cases_suite_name.append(suite_name)
    return cases_suite_name


# 需要根据要执行的测试集合名称来读取对应的测试用例数据
def read_testcases(wb, suite_name):
    sheet_data = wb[suite_name]
    lines_count = sheet_data.max_row  # 获取总行数
    cols_count = sheet_data.max_column  # 获取总列数
    """
    规定读出来的测试数据存储结构如下：
    {
        “新增客户正确”:[
            ['apiname','接口地址','请求方式','头信息',....],
            ['apiname','接口地址','请求方式','头信息',....],
        ],
        "新增客户失败-用户名为空":[
            ['apiname','接口地址','请求方式','头信息',....]
        ],
        "新增客户失败-手机号格式不正确":[
            ['apiname','接口地址','请求方式','头信息',....]
        ]
    }
    """
    cases_info = {}  # 用来存储当前测试集合中的所有用例信息的
    for l in range(2, lines_count + 1):
        case_name = sheet_data.cell(l, 2).value  # 测试用例名称
        lines = []  # 用来存储当前行测试数据的
        for c in range(3, cols_count + 1):
            cell = sheet_data.cell(l, c).value  # 当前单元格数据
            if cell == None:  # 处理空单元格
                cell = ''
            lines.append(cell)
        # 判断当前用例名称是否已存在于cases_info中
        # 如果不存在，那就是直接赋值
        # 否则就是在原来的基础上追加
        if case_name not in cases_info:
            cases_info[case_name] = [lines]
        else:
            cases_info[case_name].append(lines)
    return cases_info


# 整合所有要执行的测试用例数据，将其转成pytest参数化需要的数据结构格式
def get_all_testcases(wb):
    """
    整合后的数据结构是
    [
        ['新增客户接口测试集合','新增客户正确',[[],[]]],
        ['新增客户接口测试集合','新增客户失败-用户名为空',[[],[]]],
        ['新增客户接口测试集合','新增客户失败-手机号格式不正确',[[],[]]],
        ['新建产品接口测试集合','新建产品正确',[[],[]]],
        ['新建产品接口测试集合','新建产品失败-产品编码重复',[[],[]]],
    ]
    :param wb:
    :return:
    """
    test_data = []  # 用来存储所有测试数据
    # 获取所有要执行的测试集合名称
    cases_suite_name = get_casesuitename(wb)
    for suite_name in cases_suite_name:
        # 遍历读取每个要执行的测试集合sheet工作表中的测试用例数据
        cur_cases_info = read_testcases(wb, suite_name)  # 是个字典
        for key, value in cur_cases_info.items():
            # key实际上就是测试用例名称，value实际上测试用例多行数据信息
            case_info = [suite_name, key, value]
            test_data.append(case_info)
    return test_data


# 该方法是针对excel中数据中存在动态变量时，进行变量识别以及替换的
def regx_sub(string, vars_dict):
    res = re.findall(r'\$\{([A-Za-z_]+?)\}', string)
    for var_name in res:
        # print(var_name)
        value = vars_dict[var_name]
        # print(value)
        # 得到变量对应的值value，然后用字符串替换的方法替换
        string = string.replace(f'${{{var_name}}}', str(value))
    return string


# 针对excel数据中存在动态函数调用时，使用正则匹配并执行函数完成数据替换
def regx_func_exec(string):
    res = re.findall(r'\$\{\{(.+?)\((.*?)\)\}\}', string)
    print(res)
    for func_method in res:
        print(func_method)
        func_name = func_method[0]  # 函数名称
        func_params = func_method[1]  # 函数参数
        # 使用python中的反射机制来实现函数执行
        if hasattr(util_func, func_name):
            # 如果该函数在util_func这个文件中，那么我就去得到该函数对象
            f_method = getattr(util_func, func_name)
            if func_params == '':
                value = f_method()
                # print(value)
            else:
                value = f_method(func_params)
            string = string.replace(f'${{{{{func_name}({func_params})}}}}', str(value))
        else:
            raise BaseException(f'{func_name}不存在')
    return string


# 统一的针对数据做动态变量和动态函数的替换
def regx_sub_data(string, vars_dict):
    string = regx_sub(string, vars_dict)
    string = regx_func_exec(string)
    return string


def update_value_to_json(json_object, json_path, new_value):
    json_path_expr = parse(json_path)

    for match in json_path_expr.find(json_object):
        path = match.path  # 这是获取到匹配结果的路径
        if isinstance(path, Index):
            match.context.value[match.path.index] = new_value
        elif isinstance(path, Fields):
            match.context.value[match.path.fields[0]] = new_value
    return json_object


if __name__ == '__main__':
    wb = openpyxl.load_workbook('../testcases/CRM系统接口测试用例.xlsx')
    # print(get_variables(wb))
    # print(get_api_default_params(wb))
    # print(get_casesuitename(wb))
    # print(read_testcases(wb,'新增客户接口测试集合'))
    # print(get_all_testcases(wb))
    # s = """
    # {
    # "Admin-Token":"${token}"
    # }
    # """
    # vars_dict = {"token": "sjhdfhjdhjhjshjfdh"}
    # print(regx_sub(s, vars_dict))

    s = """
        {
        "json":{
        "$.entity.name":"联系人${{cur_timestamp()}}",
        }
    }

    """
    print(regx_func_exec(s))
